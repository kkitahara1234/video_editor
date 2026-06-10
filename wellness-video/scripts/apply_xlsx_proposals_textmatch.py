#!/usr/bin/env python3
"""
apply_xlsx_proposals_textmatch.py

xlsx の修正案を script.json にテキストマッチで反映。
cam-ID ではなく元テロップのテキストでマッチングする。

用途: xlsx と script.json の cam-ID が異なる場合（dictionary 補正後の再生成等）。

使い方:
  python3 scripts/apply_xlsx_proposals_textmatch.py \
    --xlsx <path> --script <path> --dry-run
"""

import argparse
import json
import re
import time
from pathlib import Path
from openpyxl import load_workbook


def apply_dictionary(text: str, dictionary: dict) -> str:
    """applyDictionary 同等: 長いキー優先で全件置換 + 句読点除去"""
    sorted_entries = sorted(dictionary.items(), key=lambda x: -len(x[0]))
    for from_str, to_str in sorted_entries:
        text = text.replace(from_str, to_str)
    # 句読点除去（display-corrections.ts 同等）
    text = re.sub(r'[、。]', '', text)
    # 「時→とき」補正（display-corrections.ts 同等）
    text = re.sub(r'([するたれてっいのな])時(?![間刻代計差点期分秒給空候速進制限効因報距系列節])', r'\1とき', text)
    return text


def main():
    parser = argparse.ArgumentParser(
        description='xlsx 修正案を script.json にテキストマッチで反映')
    parser.add_argument('--xlsx', required=True, help='修正案入り xlsx')
    parser.add_argument('--script', required=True, help='script.json パス')
    parser.add_argument('--dictionary', default='/Volumes/編集用/wellness-shared/dictionary.json')
    parser.add_argument('--dry-run', action='store_true', help='書き込みしない')
    args = parser.parse_args()

    # データ読み込み
    with open(args.dictionary, 'r', encoding='utf-8') as f:
        dictionary = json.load(f)

    with open(args.script, 'r', encoding='utf-8') as f:
        script = json.load(f)

    wb = load_workbook(args.xlsx, data_only=True)
    ws = wb.active

    # xlsx から修正案を抽出
    proposals = []
    for row in range(2, ws.max_row + 1):
        text_orig = ws.cell(row, 9).value   # I列: テロップ（元）
        text_fix = ws.cell(row, 10).value    # J列: 修正案
        if not text_fix or not str(text_fix).strip():
            continue
        if not text_orig:
            continue
        proposals.append((str(text_orig).strip(), str(text_fix).strip(), row))

    print(f'xlsx 修正案: {len(proposals)}件')

    # script.json の全テロップをインデックス付きで収集
    telop_index = {}  # text -> [(cam_id, idx)]
    for cam_id, telops in script['subtitles'].items():
        for idx, t in enumerate(telops):
            key = t['text']
            if key not in telop_index:
                telop_index[key] = []
            telop_index[key].append((cam_id, idx))

    # マッチング + 適用
    applied = 0
    not_found = []
    duplicates = []

    for text_orig, text_fix, row in proposals:
        # xlsx 元テロップに dictionary + fixText 適用（現 script.json と同じ補正状態にする）
        text_orig_normalized = apply_dictionary(text_orig, dictionary)

        # 修正案にも同じ補正
        text_fix_normalized = apply_dictionary(text_fix, dictionary)

        # script.json でマッチ
        matches = telop_index.get(text_orig_normalized, [])

        if len(matches) == 0:
            not_found.append((row, text_orig, text_orig_normalized))
            continue

        if len(matches) > 1:
            duplicates.append((row, text_orig, len(matches)))

        # 最初のマッチに適用
        cam_id, idx = matches[0]

        if text_orig_normalized == text_fix_normalized:
            # 補正後は同一→スキップ
            continue

        print(f'  行{row}: REPLACE')
        print(f'    - {text_orig_normalized!r}')
        print(f'    + {text_fix_normalized!r}')

        if not args.dry_run:
            script['subtitles'][cam_id][idx]['text'] = text_fix_normalized

        applied += 1

    # 結果報告
    print()
    print(f'適用: {applied}件')
    print(f'未マッチ: {len(not_found)}件')
    print(f'重複マッチ: {len(duplicates)}件')

    if not_found:
        print(f'\n--- 未マッチ詳細（先頭15件）---')
        for row, orig, normalized in not_found[:15]:
            print(f'  行{row}: 元「{orig}」')
            print(f'         補正後「{normalized}」')

    if duplicates:
        print(f'\n--- 重複マッチ ---')
        for row, orig, count in duplicates:
            print(f'  行{row}: 「{orig}」が{count}回出現')

    if not args.dry_run and applied > 0:
        # メタデータに最終手動編集タイムスタンプ
        from datetime import datetime, timezone
        if '_meta' not in script:
            script['_meta'] = {}
        script['_meta']['lastManualEditAt'] = datetime.now(timezone.utc).isoformat()
        # 書き込み
        with open(args.script, 'w', encoding='utf-8') as f:
            json.dump(script, f, ensure_ascii=False, indent=2)
        print(f'\n✅ script.json 更新済み ({applied}件)')
    elif args.dry_run:
        print(f'\n(dry-run、書き込みなし)')


if __name__ == '__main__':
    main()
