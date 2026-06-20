#!/usr/bin/env python3
"""
apply_xlsx_proposals.py — script_check.xlsx の修正案を script.json に適用

使い方:
  python scripts/apply_xlsx_proposals.py --dry-run   # diff のみ表示
  python scripts/apply_xlsx_proposals.py             # 実際に書き込み

xlsx 列構成 (make_excel.py 生成):
  B: cam_id
  C: idx
  I: テロップ（元）
  J: 修正案 ✏️
  K: アクション  (replace / delete / combine_next / 空)
"""

import json
import sys
import os
import time
import shutil
import argparse
from openpyxl import load_workbook

_parser = argparse.ArgumentParser(description='script_check.xlsx の修正案を script.json に適用')
_parser.add_argument('--dry-run', action='store_true', help='diff のみ表示、書き込みしない')
_parser.add_argument('--xlsx', default='/Volumes/編集用/script_check.xlsx', help='xlsx 入力パス')
_parser.add_argument('--script', default=None, help='script.json パス (省略時は従来の __file__ ベース)')
_args = _parser.parse_args()

SCRIPT_JSON = _args.script or os.path.join(os.path.dirname(__file__), "..", "public", "script.json")
XLSX_PATH = _args.xlsx
MAX_LEN = 25

# xlsx 列番号 (1-based)
COL_CAM_ID = 2
COL_IDX = 3
COL_ORIG_TEXT = 9
COL_PROPOSAL = 10
COL_ACTION = 11


def main():
    dry_run = _args.dry_run

    # ── xlsx 読み込み ──
    if not os.path.exists(XLSX_PATH):
        print(f"❌ {XLSX_PATH} が見つかりません")
        return 1

    wb = load_workbook(XLSX_PATH, data_only=True)
    ws = wb.active

    # ── script.json 読み込み ──
    with open(SCRIPT_JSON, "r", encoding="utf-8") as f:
        data = json.load(f)
    subtitles = data.get("subtitles", {})

    # ── xlsx から修正対象を抽出 ──
    proposals = []
    warnings = []

    for row_num in range(2, ws.max_row + 1):
        proposal = str(ws.cell(row=row_num, column=COL_PROPOSAL).value or "").strip()
        action = str(ws.cell(row=row_num, column=COL_ACTION).value or "").strip().lower()

        # 空行はスキップ
        if not proposal and not action:
            continue

        cam_id = str(ws.cell(row=row_num, column=COL_CAM_ID).value or "").strip()
        idx_raw = ws.cell(row=row_num, column=COL_IDX).value
        orig_text = str(ws.cell(row=row_num, column=COL_ORIG_TEXT).value or "").strip()

        # idx を整数化
        try:
            idx = int(idx_raw)
        except (TypeError, ValueError):
            warnings.append(f"  行{row_num}: idx が不正 ({idx_raw!r})")
            continue

        # アクション判定
        if not action and proposal:
            action = "replace"  # デフォルト

        if action not in ("replace", "delete", "combine_next"):
            warnings.append(f"  行{row_num} {cam_id}[{idx}]: 不正なアクション {action!r}")
            continue

        # cam_id 存在チェック
        if cam_id not in subtitles:
            warnings.append(f"  行{row_num}: cam_id {cam_id!r} が subtitles に不在")
            continue

        arr = subtitles[cam_id]

        # idx 範囲チェック
        if idx < 0 or idx >= len(arr):
            warnings.append(f"  行{row_num} {cam_id}[{idx}]: idx 範囲外 (len={len(arr)})")
            continue

        # 元テキスト一致チェック
        current_text = arr[idx]["text"]
        if current_text != orig_text:
            warnings.append(
                f"  行{row_num} {cam_id}[{idx}]: 元テキスト不一致\n"
                f"    xlsx:   {orig_text!r}\n"
                f"    json:   {current_text!r}"
            )
            continue

        proposals.append({
            "row": row_num,
            "cam_id": cam_id,
            "idx": idx,
            "action": action,
            "proposal": proposal,
            "orig_text": orig_text,
        })

    if not proposals and not warnings:
        print("修正対象なし（修正案・アクション列は全て空）")
        return 0

    # ── アクション別バリデーション ──
    valid = []
    for p in proposals:
        cam_id, idx, action, proposal = p["cam_id"], p["idx"], p["action"], p["proposal"]
        arr = subtitles[cam_id]

        if action == "replace":
            if not proposal:
                warnings.append(f"  行{p['row']} {cam_id}[{idx}]: replace だが修正案が空")
                continue
            if len(proposal) > MAX_LEN:
                warnings.append(
                    f"  行{p['row']} {cam_id}[{idx}]: 修正案が {len(proposal)}字 (>{MAX_LEN}字)\n"
                    f"    修正案: {proposal!r}"
                )
                continue

        elif action == "delete":
            if len(arr) <= 1:
                warnings.append(f"  行{p['row']} {cam_id}[{idx}]: 最後の1件は削除不可")
                continue

        elif action == "combine_next":
            if idx + 1 >= len(arr):
                warnings.append(f"  行{p['row']} {cam_id}[{idx}]: 次テロップが存在しない")
                continue
            combined = proposal if proposal else (arr[idx]["text"] + arr[idx + 1]["text"])
            if len(combined) > MAX_LEN:
                warnings.append(
                    f"  行{p['row']} {cam_id}[{idx}]: 結合後 {len(combined)}字 (>{MAX_LEN}字)\n"
                    f"    結合後: {combined!r}"
                )
                continue
            p["_combined_text"] = combined

        valid.append(p)

    # ── アクション別集計 ──
    counts = {"replace": 0, "delete": 0, "combine_next": 0}
    for p in valid:
        counts[p["action"]] += 1

    mode = "[DRY-RUN] " if dry_run else ""
    print("=" * 60)
    print(f"{mode}修正対象: {len(valid)}件")
    print("=" * 60)
    for a, c in counts.items():
        if c > 0:
            print(f"  {a}: {c}件")
    print()

    # ── diff 表示 ──
    for p in valid:
        cam_id, idx, action = p["cam_id"], p["idx"], p["action"]
        arr = subtitles[cam_id]

        if action == "replace":
            print(f"  {cam_id}[{idx}]: REPLACE")
            print(f"    - {p['orig_text']!r}")
            print(f"    + {p['proposal']!r}")

        elif action == "delete":
            print(f"  {cam_id}[{idx}]: DELETE")
            print(f"    - {p['orig_text']!r}")

        elif action == "combine_next":
            next_text = arr[idx + 1]["text"]
            print(f"  {cam_id}[{idx}]+[{idx+1}]: COMBINE_NEXT")
            print(f"    - {p['orig_text']!r} + {next_text!r}")
            print(f"    + {p['_combined_text']!r}")

    # ── 警告 ──
    if warnings:
        print(f"\n⚠ 警告: {len(warnings)}件")
        for w in warnings:
            print(w)

    # ── 本番適用 ──
    if dry_run:
        return 0 if not warnings else 1

    if not valid:
        print("\n適用可能な修正がありません")
        return 1

    # バックアップ
    bak = f"{SCRIPT_JSON}.bak.{int(time.time())}"
    with open(bak, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
        f.write("\n")
    print(f"\nバックアップ: {bak}")

    # インデックスずれ防止: cam_id ごとに idx 降順でソート
    valid.sort(key=lambda p: (p["cam_id"], -p["idx"]))

    applied = 0
    for p in valid:
        cam_id, idx, action = p["cam_id"], p["idx"], p["action"]
        arr = subtitles[cam_id]

        if action == "replace":
            arr[idx]["text"] = p["proposal"]
            applied += 1

        elif action == "delete":
            arr.pop(idx)
            applied += 1

        elif action == "combine_next":
            combined_text = p["_combined_text"]
            next_sub = arr[idx + 1]
            arr[idx]["text"] = combined_text
            arr[idx]["absEndSec"] = next_sub["absEndSec"]
            arr[idx]["durationSec"] = next_sub["absEndSec"] - arr[idx]["absStartSec"]
            if "endSec" in next_sub:
                arr[idx]["endSec"] = next_sub["endSec"]
            arr.pop(idx + 1)
            applied += 1

    # 書き込み
    with open(SCRIPT_JSON, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
        f.write("\n")

    print(f"\n✓ 適用: {applied}件")
    print(f"✓ {SCRIPT_JSON} を更新しました")

    # ── --resync: 適用後にタイミング再同期 DRY-RUN ──
    if "--resync" in sys.argv:
        print()
        print("=" * 60)
        print("タイミング再同期 (自動呼び出し)")
        print("=" * 60)
        import subprocess
        resync_script = os.path.join(os.path.dirname(__file__), "auto_resync_timing.py")
        cmd = [sys.executable, resync_script, "--dry-run", "--backup-path", bak]
        subprocess.run(cmd)

    return 0 if not warnings else 1


if __name__ == "__main__":
    sys.exit(main())
