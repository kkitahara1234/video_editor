#!/usr/bin/env python3
"""
whisperx_to_xlsx.py
whisperX の master.json → script_check.xlsx 変換

- word 累積文字数で18字前後に行分割（助詞直後・文字種境界を優先）
- 句読点（、。！？）は除去（既存パイプライン準拠）
- 話者(host/guest)を cam_id / angle 列に格納
- 時刻は word の実時刻を使用（文字数按分ではない）
- xlsx-loader.ts が読める subtitles シート形式で出力
"""

import json
import re
import argparse
from collections import Counter
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment

# ── 分割パラメータ ──
SOFT_MAX = 18      # この文字数を超えたら分割候補を探す
HARD_MAX = 25      # 強制分割上限
MIN_LINE = 4       # 分割時の最小行文字数

PUNCT = re.compile(r'[、。！？,.?!]')

# 助詞（この直後で切りやすい）
JOSHI = set('をがにはでともやかのへ')
# 終助詞的（この直後は切らない — 次の文頭にすべき）
JOSHI_EXCLUDE_NEXT = set('ねよぞさわ')


def fmt_time(sec: float) -> str:
    h = int(sec // 3600)
    m = int((sec % 3600) // 60)
    s = sec % 60
    return f"{h:02d}:{m:02d}:{s:06.3f}"


def is_hiragana(c: str) -> bool:
    return '\u3040' <= c <= '\u309f'

def is_katakana(c: str) -> bool:
    return '\u30a0' <= c <= '\u30ff'

def is_kanji(c: str) -> bool:
    return '\u4e00' <= c <= '\u9fff'

def is_char_boundary(prev: str, curr: str) -> bool:
    """文字種が変わる境界か（ひらがな↔漢字, ひらがな↔カタカナ）"""
    if not prev or not curr:
        return False
    hp, hc = is_hiragana(prev), is_hiragana(curr)
    kp, kc = is_kanji(prev), is_kanji(curr)
    tp, tc = is_katakana(prev), is_katakana(curr)
    return (hp and kc) or (kp and hc) or (hp and tc) or (tp and hc)


def find_best_split(text: str, soft: int = SOFT_MAX, hard: int = HARD_MAX) -> int:
    """text 中の分割位置を返す。見つからなければ hard で強制切り。"""
    if len(text) <= soft:
        return -1  # 分割不要

    min_pos = max(MIN_LINE, int(soft * 0.4))  # ≈ 7
    max_pos = soft

    # ① 助詞の直後（範囲内で最後のもの）
    best = -1
    for pos in range(min_pos, min(max_pos + 1, len(text))):
        prev = text[pos - 1]
        nxt = text[pos] if pos < len(text) else ''
        if prev in JOSHI and nxt not in JOSHI_EXCLUDE_NEXT:
            # カタカナ連続・漢字連続の途中は避ける
            if not (is_katakana(prev) and is_katakana(nxt)):
                if not (is_kanji(prev) and is_kanji(nxt)):
                    best = pos
    if best != -1:
        return best

    # ② 文字種境界（範囲内で最後のもの）
    for pos in range(max_pos, min_pos - 1, -1):
        if pos <= 0 or pos >= len(text):
            continue
        if is_char_boundary(text[pos - 1], text[pos]):
            return pos

    # ③ 接続助詞「て」の直後
    for pos in range(max_pos, min_pos - 1, -1):
        if pos <= 0 or pos >= len(text):
            continue
        if text[pos - 1] == 'て':
            return pos

    # ④ 強制: hard で切る
    return min(hard, len(text))


def split_segment(seg: dict) -> list[dict]:
    """segment の words を行単位に分割して返す"""
    words = seg.get('words', [])
    speaker = seg.get('speaker', 'unknown')

    if not words:
        text = PUNCT.sub('', seg.get('text', '')).strip()
        if not text:
            return []
        return [{'text': text, 'start': seg['start'], 'end': seg['end'], 'speaker': speaker}]

    # ── word を順に累積し、クリーンテキスト + word index のマッピングを作る ──
    # char_map[i] = その文字を含む word の index
    clean_chars: list[str] = []       # 句読点除去後の文字列
    char_to_word: list[int] = []      # clean_chars[i] が由来する words[] の index

    for wi, w in enumerate(words):
        raw = w.get('word', '')
        for ch in raw:
            if not PUNCT.match(ch):
                clean_chars.append(ch)
                char_to_word.append(wi)

    full_text = ''.join(clean_chars)
    if not full_text.strip():
        return []

    # ── full_text を再帰的に分割位置で切る → 各チャンクの文字範囲を決定 ──
    def split_recursive(start: int, end: int) -> list[tuple[int, int]]:
        """clean_chars[start:end] を分割し、(start, end) ペアのリストを返す"""
        length = end - start
        if length <= SOFT_MAX:
            return [(start, end)]
        text = ''.join(clean_chars[start:end])
        sp = find_best_split(text)
        if sp == -1 or sp <= 0 or sp >= length:
            return [(start, end)]
        abs_sp = start + sp
        return split_recursive(start, abs_sp) + split_recursive(abs_sp, end)

    chunks = split_recursive(0, len(clean_chars))

    # ── 各チャンクを line に変換（word 実時刻を使用）──
    lines = []
    for c_start, c_end in chunks:
        if c_start >= c_end:
            continue
        text = ''.join(clean_chars[c_start:c_end]).strip()
        if not text:
            continue

        # このチャンクに含まれる word の index 範囲
        first_wi = char_to_word[c_start]
        last_wi = char_to_word[c_end - 1]

        lines.append({
            'text': text,
            'start': words[first_wi].get('start', seg['start']),
            'end':   words[last_wi].get('end', seg['end']),
            'speaker': speaker,
        })

    return lines


def main():
    parser = argparse.ArgumentParser(description='whisperX master.json → script_check.xlsx')
    parser.add_argument('--master', default='public/master.json', help='master.json パス')
    parser.add_argument('--xlsx',   default='script_check.xlsx',  help='xlsx 出力先')
    args = parser.parse_args()

    with open(args.master, 'r', encoding='utf-8') as f:
        master = json.load(f)

    # ── 全セグメント → 行分割 ──
    all_lines = []
    for seg in master['segments']:
        all_lines.extend(split_segment(seg))

    print(f"セグメント数: {len(master['segments'])}")
    print(f"分割後テロップ行数: {len(all_lines)}")

    # ── xlsx 書き出し ──
    wb = Workbook()
    ws = wb.active
    ws.title = "subtitles"

    headers = ['#', 'cam_id', 'idx', 'angle', 'topic', '絶対開始', '絶対終了', '秒',
               'テロップ（元）', '修正案 ✏️', 'アクション', '文字数', '⚠️']
    ws.append(headers)

    # ヘッダスタイル
    header_blue   = PatternFill('solid', fgColor='305496')
    header_yellow = PatternFill('solid', fgColor='BF8F00')
    for c in ws[1]:
        c.font = Font(bold=True, color='FFFFFF')
        c.fill = header_blue
        c.alignment = Alignment(horizontal='center')
    ws.cell(row=1, column=10).fill = header_yellow
    ws.cell(row=1, column=11).fill = header_yellow

    # 行色定義
    red    = PatternFill('solid', fgColor='F8CBAD')
    yellow = PatternFill('solid', fgColor='FFF2CC')
    orange = PatternFill('solid', fgColor='FFD966')
    edit_bg = PatternFill('solid', fgColor='FFFDE7')

    # 話者ごとの idx カウンタ
    spk_idx: dict[str, int] = {}

    for i, line in enumerate(all_lines, 1):
        spk = line['speaker']
        if spk not in spk_idx:
            spk_idx[spk] = 0
        idx = spk_idx[spk]
        spk_idx[spk] += 1

        dur  = round(line['end'] - line['start'], 3)
        text = line['text']
        n    = len(text)

        # フラグ
        if n <= 1:    flag = '🔴1文字'
        elif n <= 3:  flag = '⚠️短い'
        elif n >= 35: flag = '🟠長い'
        else:         flag = ''

        row_num = i + 1  # Excel行（ヘッダが1行目）
        ws.append([
            i,                       # A: #
            spk,                     # B: cam_id (= speaker)
            idx,                     # C: idx
            spk,                     # D: angle (= speaker → カメラ切替用)
            '',                      # E: topic (空。step1 で GPT が埋める)
            fmt_time(line['start']), # F: 絶対開始
            fmt_time(line['end']),   # G: 絶対終了
            dur,                     # H: 秒
            text,                    # I: テロップ（元）
            None,                    # J: 修正案 ✏️
            None,                    # K: アクション
            n,                       # L: 文字数
            flag if flag else None,  # M: ⚠️
        ])

        # J,K列の背景（編集欄）
        ws.cell(row=row_num, column=10).fill = edit_bg
        ws.cell(row=row_num, column=11).fill = edit_bg

        # フラグ行の色付け
        if flag.startswith('🔴'):
            for col in range(1, 14):
                ws.cell(row=row_num, column=col).fill = red
        elif flag.startswith('⚠️'):
            for col in range(1, 14):
                ws.cell(row=row_num, column=col).fill = yellow
        elif flag.startswith('🟠'):
            for col in range(1, 14):
                ws.cell(row=row_num, column=col).fill = orange

    # 列幅
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 5
    ws.column_dimensions['D'].width = 8
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 16
    ws.column_dimensions['G'].width = 16
    ws.column_dimensions['H'].width = 8
    ws.column_dimensions['I'].width = 40
    ws.column_dimensions['J'].width = 40
    ws.column_dimensions['K'].width = 12
    ws.column_dimensions['L'].width = 6
    ws.column_dimensions['M'].width = 12

    wb.save(args.xlsx)
    print(f"xlsx 書き出し: {args.xlsx}")

    # ── 統計 ──
    spk_count = Counter(l['speaker'] for l in all_lines)
    lens = [len(l['text']) for l in all_lines]
    print(f"話者別行数: {dict(spk_count)}")
    print(f"文字数: 最小{min(lens)} 最大{max(lens)} 平均{sum(lens)//len(lens)}")

    # 文字数分布
    bins = [(0, 5), (5, 10), (10, 15), (15, 20), (20, 25), (25, 30), (30, 40), (40, 999)]
    for lo, hi in bins:
        cnt = sum(1 for l in lens if lo <= l < hi)
        if cnt:
            print(f"  {lo:2d}-{hi:2d}字: {cnt}件")


if __name__ == '__main__':
    main()
