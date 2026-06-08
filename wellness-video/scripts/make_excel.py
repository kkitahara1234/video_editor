import json, os, re
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment

with open('public/script.json','r',encoding='utf-8') as f:
    data = json.load(f)

subs = data['subtitles']
segments = {s['id']: s for s in data['segments']}

def fmt_time(sec):
    try:
        sec = float(sec)
        h = int(sec // 3600); m = int((sec % 3600) // 60); s = sec % 60
        return f"{h:02d}:{m:02d}:{s:06.3f}"
    except: return ""

def has_content(text):
    if re.search(r'[\u4E00-\u9FFF]', text): return True
    if re.search(r'[\u30A0-\u30FF]{2,}', text): return True
    if re.search(r'[A-Za-z0-9]', text): return True
    return False

rows = []
for cam_id in sorted(subs.keys()):
    seg = segments.get(cam_id, {})
    topic = seg.get('topicLabel', '')
    angle = seg.get('angle', '')
    for idx, t in enumerate(subs[cam_id]):
        text = t.get('text','')
        n = len(text.strip())
        no_content = not has_content(text)
        if n <= 1: flag = '🔴1文字'
        elif n <= 3: flag = '⚠️短い'
        elif no_content: flag = '🟡自立語なし'
        elif n >= 22: flag = '🟠長い'
        else: flag = ''
        rows.append({
            'cam_id': cam_id, 'idx_in_cam': idx, 'angle': angle, 'topic': topic,
            'absStart': fmt_time(t.get('absStartSec', 0)),
            'absEnd': fmt_time(t.get('absEndSec', 0)),
            'duration': round(float(t.get('durationSec', 0)), 2),
            'text': text, 'len': n, 'flag': flag, 'no_content': no_content,
        })

print(f"集計テロップ数: {len(rows)}")

wb = Workbook(); ws = wb.active; ws.title = "subtitles"
# ── 列構成: A-M (13列) ──
headers = ['#','cam_id','idx','angle','topic','絶対開始','絶対終了','秒',
           'テロップ（元）','修正案 ✏️','アクション','文字数','⚠️']
ws.append(headers)

# ── ヘッダーのスタイル ──
header_blue   = PatternFill('solid', fgColor='305496')
header_yellow = PatternFill('solid', fgColor='BF8F00')  # 濃い黄 (白文字が映える)

for c in ws[1]:
    c.font = Font(bold=True, color='FFFFFF')
    c.fill = header_blue
    c.alignment = Alignment(horizontal='center')
# J列 (修正案) と K列 (アクション) は黄色ヘッダー
ws.cell(row=1, column=10).fill = header_yellow
ws.cell(row=1, column=11).fill = header_yellow

# ── 行の色定義 ──
red    = PatternFill('solid', fgColor='F8CBAD')
yellow = PatternFill('solid', fgColor='FFF2CC')
gold   = PatternFill('solid', fgColor='FFE699')
orange = PatternFill('solid', fgColor='FFD966')
gray   = PatternFill('solid', fgColor='F2F2F2')
edit_bg = PatternFill('solid', fgColor='FFFDE7')  # 薄黄: 記入欄

prev_cam = None
for i, r in enumerate(rows, 1):
    ws.append([i, r['cam_id'], r['idx_in_cam'], r['angle'], r['topic'],
               r['absStart'], r['absEnd'], r['duration'], r['text'],
               '', '',  # J:修正案, K:アクション (空)
               r['len'], r['flag']])
    row_n = ws.max_row

    # 行全体の色付け (フラグ行)
    if r['len'] <= 1:
        for c in ws[row_n]: c.fill = red
    elif r['len'] <= 3:
        for c in ws[row_n]: c.fill = yellow
    elif r['no_content']:
        for c in ws[row_n]: c.fill = gold
    elif r['len'] >= 22:
        for c in ws[row_n]: c.fill = orange
    elif r['cam_id'] != prev_cam and i % 2 == 0:
        for c in ws[row_n]: c.fill = gray

    # J列 (修正案) と K列 (アクション) は常に薄黄背景
    ws.cell(row=row_n, column=10).fill = edit_bg
    ws.cell(row=row_n, column=11).fill = edit_bg
    prev_cam = r['cam_id']

# ── 列幅 ──
widths = {'A':6,'B':12,'C':6,'D':10,'E':22,'F':14,'G':14,'H':8,
          'I':50,'J':50,'K':16,'L':8,'M':14}
for col, w in widths.items():
    ws.column_dimensions[col].width = w

# ── 表示設定 ──
ws.freeze_panes = 'A2'
ws.auto_filter.ref = ws.dimensions
for row in ws.iter_rows(min_row=2, min_col=9, max_col=11):
    for cell in row:
        cell.alignment = Alignment(wrap_text=True, vertical='top')

out = '/Volumes/編集用/script_check.xlsx'
wb.save(out)

print(f"✅ 保存: {out}")
print(f"  総数: {len(rows)}件")
print(f"  🔴 1文字: {sum(1 for r in rows if r['len']<=1)}件")
print(f"  ⚠️ 短い: {sum(1 for r in rows if 1<r['len']<=3)}件")
print(f"  🟡 自立語なし: {sum(1 for r in rows if r['no_content'] and r['len']>3)}件")
print(f"  🟠 22字以上: {sum(1 for r in rows if r['len']>=22)}件")
